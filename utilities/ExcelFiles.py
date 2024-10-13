import openpyxl


# briefing the values from excell file
def get_data_from_excell(path, sheet_name):
    final_list = []
    work_book = openpyxl.load_workbook(path)    # read excell file
    sheet = work_book[sheet_name]   # select sheet
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            row_list.append(sheet.cell(row=r, column=c).value)  # get values of rows and columns from rows into row_list
        final_list.append(row_list)
    return final_list
