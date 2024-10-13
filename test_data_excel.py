import pytest
import openpyxl


def get_data():
    final_list = []
    work_book = openpyxl.load_workbook('ExcelFiles/loginfile.xlsx')
    sheet = work_book['LoginPage']
    total_rows = sheet.max_row
    total_cols = sheet.max_column
    for r in range(2, total_rows+1):
        row_list = []
        for c in range(1, total_cols+1):
            row_list.append(sheet.cell(row=r,column=c).value)
        final_list.append(row_list)
    return final_list


@pytest.mark.parametrize("email_address, password", get_data())
def test_email_password(email_address, password):
    print(f"My email is: {email_address} and password is: {str(password)}")
